import pandas as pd
import time
import datetime
import wget
from os import scandir, getcwd
import openpyxl
# import git
import requests
from os import remove

def  mensaje():
    print("Este es un mensaje dentro de una función.")

def  mensaje2():
    print("Este es un mensaje dentro de una función 2.")

def Ciclo():
    Archivos = Descargar_Archivos()
    Archivos = lsExcel()
    if(len(Archivos) > 0):
        Actualizar_Datos(Archivos)
        # guardarRepositorio()
        # time.sleep(60 * 60 * 24)
    else:
        print("No hay datos que actualizar")
        #time.sleep(60 * 60 * 4)
    print("Ciclo completo")
    # Ciclo()

def Descargar_Archivos():
    fechaMaxima = Fecha_Actual_Fruta()
    salida = []
    numero = 1
    # Flag = Verificar si la fecha máxima del archivo es diferente a la fecha actual.
    flag = (fechaMaxima + datetime.timedelta(days = 0)).strftime("%Y%m%d") != datetime.datetime.now().strftime("%Y%m%d")
    print(flag)

    while flag:
        # annioDescarga = None
        annioDescarga = (fechaMaxima + datetime.timedelta(days = numero)).strftime("%Y")
        mesDescarga = (fechaMaxima + datetime.timedelta(days = numero)).strftime("%m")
        urlBase =  "https://www.odepa.gob.cl/wp-content/uploads/" + annioDescarga + "/" + mesDescarga + "/Boletin_Diario_de_Frutas_y_Hortalizas_"
        fecha = (fechaMaxima + datetime.timedelta(days = numero)).strftime("%Y%m%d.xlsx") 
        # print(urlBase + fecha)

        try:
            myfile = requests.get(urlBase + fecha)
            open(fecha, 'wb').write(myfile.content)
            #wget.download(urlBase + fecha, fecha)
            print(urlBase + fecha)
            #salida.append(fecha)
        except:
            pass
        
        try:
            pd.read_excel(fecha)
            salida.append(fecha)
        except:
            remove(fecha)

        flag = (fechaMaxima + datetime.timedelta(days = numero)).strftime("%Y%m%d") != datetime.datetime.now().strftime("%Y%m%d")
        numero = numero + 1
    
    return salida

def lsExcel():
    salida = []
    for i in ls():
        if("xlsx" in i and "20" in i and ".tmp" not in i and i[0] == "2"):
            salida.append(i)
    return salida

def SalidaFecha(nombre):
    #20201124
    #01234567
    #27-10-2020
    #return nombre[6:8] + "-" + nombre[4:6] + "-" + nombre[0:4]
    fecha_str = nombre[6:8] + "-" + nombre[4:6] + "-" + nombre[0:4]
    return datetime.datetime.strptime(fecha_str,"%d-%m-%Y")

def diccionario_auxiliar(Mercado, Region, Fecha, Codreg, Tipo, Categoria, Producto,
       Variedad, Calidad, Volumen, Precio_minimo, Precio_maximo,
       Precio_promedio_ponderado, Unidad_de_comercializacion, Origen,
       Precio_Kg, Kg_unidad):
    return {'Mercado' : Mercado, 
        'Región' : Region, 
        'Fecha' : Fecha, 
        'Codreg' : Codreg, 
        'Tipo' : Tipo, 
        'Categoría' : Categoria, 
        'Producto' : Producto,
        'Variedad' : Variedad, 
        'Calidad' : Calidad, 
        'Volumen' : Volumen, 
        'Precio mínimo' : Precio_minimo, 
        'Precio máximo' : Precio_maximo,
       'Precio promedio ponderado' : Precio_promedio_ponderado, 
        'Unidad de comercialización' : Unidad_de_comercializacion, 
        'Origen':Origen,
       'Precio $/Kg' : Precio_Kg, 
        'Kg / unidad' : Kg_unidad}

def convertirFecha(fecha):
    if(type(fecha) == str):
        return datetime.datetime.strptime(string,"%d-%m-%Y")
    return fecha

def Actualizar_Datos(Archivos):
    wb = openpyxl.load_workbook("Diccionario.xlsx")
    hojas_for_dict = wb.sheetnames
    hojas_for_dict

    Mercado = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[0])
    Mercado.to_dict(orient = "list")["Mercado  "]
    for i in range(len(Mercado.to_dict(orient = "list")["Mercado  "])):
        print('"' + Mercado.to_dict(orient = "list")["Mercado  "][i] + '"')
        
    Mercado_Dict = {
        'Lo Valledor':["Mercado Mayorista Lo Valledor de Santiago",13],
    'Vega Central Mapocho':["Vega Central Mapocho de Santiago",13],
    'Macroferia Talca':["Macroferia Regional de Talca",7],
    'Femacal':["Femacal de La Calera",5],
    'La Palmera':["Terminal La Palmera de La Serena",4],
    'Solcoagro':["Comercializadora del Agro de Limarí",4],
    'Vega Monumental':["Vega Monumental Concepción",8],
    'Lagunita Pto.Montt':["Feria Lagunitas de Puerto Montt",10],
    'Vega Modelo Temuco':["Vega Modelo de Temuco",9],
    'Agrochillan':["Terminal Hortofrutícola Agro Chillán",16],
    'Agronor':["Agrícola del Norte S.A. de Arica",15],
    'Mapocho Vta.dir':  ["Mapocho Venta Directa de Santiago",13]
    }

    Region_Dict = {
        13:"Metropolitana",
        15:"Arica y Parinacota",
        4:"Coquimbo",
        5:"Coquimbo",
        10:"Los Lagos",
        7:"Maule",
        16:"Ñuble",
        9:"La Araucanía",
        8:"Bíobío"
    }

    Mes = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[1])

    Mes_Dict = {}
    for i in range(len(Mes)):
        Mes_Dict[i] = Mes["Mes"][i]
    Mes_Dict

    Especie  = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[2])
    Especie_Dict = {}
    for i in range(len(Especie)):
        Especie_Dict[Especie["Especie"][i]] = Especie["Clasificación"][i]
        
    Detalle  = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[3])
    Detalle_Dict = {}
    for i in range(len(Detalle)):
        Detalle_Dict[Detalle["Detalle"][i]] = Detalle["Kg"][i]
    Detalle_Dict['$/bandeja 18 kilos empedrada'] = 18
    Detalle_Dict['$/caja 18 kilos importada'] = 18
    Detalle_Dict['$/malla 22 kilos'] = 22
    Detalle_Dict['$/paquete 2 kilos'] = 2
    Detalle_Dict['$/caja 5 kilos'] = 5
    Detalle_Dict['$/atado'] = 1   #Preguntar
    Detalle_Dict['$/caja 8 kilos'] = 8
    Detalle_Dict['$/malla 100 unidades'] = 10 #Preguntar
    Detalle_Dict['$/media docena de atados'] = 6 #PReguntar
    Detalle_Dict['$/bins (500 kilos)'] = 500
    Detalle_Dict['$/envase 1 kilo'] = 1
    Detalle_Dict['$/cien'] = 10 #Preguntar
    Detalle_Dict['$/docena'] = 1.2 #Preguntar
    Detalle_Dict['$/caja 14 kilos'] = 14 #Preguntar
    Detalle_Dict['$/cien en rama (volumen en unidades)'] = 1 #Preguntar
    Detalle_Dict['$/caja 20 kilos empedrada'] = 20
    Detalle_Dict['$/caja 17 kilos empedrada'] = 17

    Detalle  = pd.read_excel("Diccionario.xlsx", sheet_name=hojas_for_dict[3])
    Frutas = []
    Hortalizas = []
    for i in Archivos:
        print(i)
        wb = openpyxl.load_workbook(i)
        hojas = wb.sheetnames
        hojas
        
        dict_auxiliar = {}
        for hoja in hojas:
            if("Frutas" in hoja):
                #Frutas.append(hoja)
                datos = pd.read_excel(i, sheet_name=hoja, skiprows=8, skipfooter=1)
                mercado_list = hoja.split("_")[1]
                mercado = Mercado_Dict[mercado_list][0]
                region = Region_Dict[Mercado_Dict[mercado_list][1]]
                cod_reg = Mercado_Dict[mercado_list][1]
                fecha = SalidaFecha(i)
                tipo = "Fruta"
                #print(mercado,region,fecha, cod_reg)
                #print(len(datos))
                for filas in range(len(datos)):
                    categoria = Especie_Dict[datos["Producto "][filas]]
                    producto = datos["Producto "][filas]
                    variedad = datos['Variedad '][filas]
                    calidad = datos['Calidad '][filas]
                    volumen = datos['Volumen '][filas]
                    precio_minimo = datos['Precio\nmínimo'][filas]
                    precio_maximo = datos['Precio\nmáximo'][filas]
                    precio_promedio = datos['Precio\npromedio'][filas]
                    u_comercializacion = datos['Unidad de\ncomercialización '][filas]
                    origen = datos['Origen '][filas]
                    try:
                        kgUnidad = Detalle_Dict[u_comercializacion]
                    except:
                        kgUnidad = 1
                    
                    
                    precio = int(round(precio_promedio / kgUnidad,0))               
                    Frutas.append(diccionario_auxiliar(mercado,region,fecha,cod_reg,tipo,categoria,producto,variedad,calidad,volumen,precio_minimo,precio_maximo,precio_promedio,u_comercializacion,origen, precio,kgUnidad))
                    #Frutas.append(
                #break
        
        for hoja in hojas:
            if("Hortalizas" in hoja):
                #Frutas.append(hoja)
                datos = pd.read_excel(i, sheet_name=hoja, skiprows=8, skipfooter=1)
                mercado_list = hoja.split("_")[1]
                mercado = Mercado_Dict[mercado_list][0]
                region = Region_Dict[Mercado_Dict[mercado_list][1]]
                cod_reg = Mercado_Dict[mercado_list][1]
                fecha = SalidaFecha(i)
                tipo = "Fruta"
                #print(mercado,region,fecha, cod_reg)
                #print(len(datos))
                for filas in range(len(datos)):
                    categoria = ""
                    producto = datos["Producto "][filas]
                    variedad = datos['Variedad '][filas]
                    calidad = datos['Calidad '][filas]
                    volumen = datos['Volumen '][filas]
                    precio_minimo = datos['Precio\nmínimo'][filas]
                    precio_maximo = datos['Precio\nmáximo'][filas]
                    precio_promedio = datos['Precio\npromedio'][filas]
                    try:
                        u_comercializacion = datos['Unidad de\ncomercialización '][filas]
                    except:
                        u_comercializacion = datos['Unidad de\ncomercialización'][filas]
                    origen = datos['Origen '][filas]
                    try:
                        kgUnidad = Detalle_Dict[u_comercializacion]
                    except:
                        kgUnidad = 1
                    #kgUnidad = Detalle_Dict[u_comercializacion]
                    precio = int(round(precio_promedio / kgUnidad,0))               
                    Hortalizas.append(diccionario_auxiliar(mercado,region,fecha,cod_reg,tipo,categoria,producto,variedad,calidad,volumen,precio_minimo,precio_maximo,precio_promedio,u_comercializacion,origen, precio,kgUnidad))
    datosFruta = pd.DataFrame(Frutas)
    datosHortaliza = pd.DataFrame(Hortalizas)
    
    datosHortaliza["Kg o Unidades"] = datosHortaliza["Kg / unidad"]
    del datosHortaliza["Tipo"]
    del datosHortaliza["Categoría"]
    del datosHortaliza["Kg / unidad"]

    fruta_salida = pd.concat([ref_frutas1(),datosFruta])
    hortaliza_salida = pd.concat([ref_hortalizas1(),datosHortaliza])
    fruta_salida.fillna(0)
    hortaliza_salida.fillna(0)
    #fruta_salida = datosFruta
    #hortaliza_salida = datosHortaliza
    hortaliza_salida["Clasificación"] = "Hortaliza"
    fruta_salida.to_excel("Consolidado/FrutaConsolidado.xlsx", index=False)
    hortaliza_salida.to_excel("Consolidado/HortalizaConsolidado.xlsx", index=False)
    return

def Fecha_Actual_Fruta():
    return max(ref_frutas()["Fecha"])

def ref_frutas():
    dataReferenciaFruta = pd.read_excel("Consolidado/FrutaConsolidado.xlsx")
    return dataReferenciaFruta

def ls(ruta = getcwd()):
    return [arch.name for arch in scandir(ruta) if arch.is_file()]

def ref_hortalizas():
    dataReferenciaHortaliza = pd.read_excel("Consolidado/HortalizaConsolidado.xlsx")
    return dataReferenciaHortaliza

def ref_frutas1():
    dataReferenciaFruta = pd.read_excel("Consolidado/FrutaConsolidado1.xlsx")
    return dataReferenciaFruta

def ref_hortalizas1():
    dataReferenciaHortaliza = pd.read_excel("Consolidado/HortalizaConsolidado1.xlsx")
    return dataReferenciaHortaliza

def Fecha_Actual_Fruta():
    return max(ref_frutas()["Fecha"])

def Fecha_Actual_Hortaliza():
    return max(ref_hortalizas()["Fecha"])


if __name__ == '__main__':
    print('El proceso ha comenzado.')
    Ciclo()
    print('El proceso ha finalizado.')